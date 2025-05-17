"""
Slackから投稿を取得する

create 2025/05/09 hamada
"""
import os
import sqlite3
import sys
import datetime
import time
import json

from tkinter import messagebox
import pandas as pd
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment, PatternFill
from openpyxl.formatting.rule import FormulaRule
from openpyxl.drawing.image import Image

from common import const, shared_service
from common.logger.logger import Logger
from app_common import app_shared_service
from app_common.slack_service import SlackService
from dataaccess.common import set_cond_model, set_sort_model
from dataaccess.ext.slack_export_dataaccess import SlackExportDataaccess
from dataaccess.general.channel_dataaccess import ChannelDataAccess
from dataaccess.general.slack_user_dataaccess import SlackUserDataAccess
from dataaccess.general.tr_channel_histories_dataaccess import TrChannelHistoriesDataAccess

pd.options.display.float_format = '{:.6f}'.format


class GetMessages:
    def __init__(self, logger:Logger, root_dir, bin_dir, conn, is_json=False):
        self.logger = logger
        self.root_dir = root_dir
        self.bin_dir = bin_dir
        self.conn = conn
        self.is_json = is_json

        # dataaccess
        self.channel_dataaccess = ChannelDataAccess(conn)
        self.slack_user_dataaccess = SlackUserDataAccess(conn)
        self.tr_channel_histories_dataaccess = TrChannelHistoriesDataAccess(conn)

        # 設定ファイル
        self.conf = app_shared_service.get_conf(self.root_dir)['get_messages']
        # SlackAPIトークンを取得
        token = app_shared_service.get_token()
        # Slackサービス
        if not is_json:
            if int(self.conf['is_get_messages']):
                if token is None:
                    self.logger.error('Not Define ENVIRONMENT_KEY [SLACK_API_TOKEN]')
                    messagebox.showwarning('WARN', '環境変数「SLACK_API_TOKEN」が定義されていません')
                    sys.exit()
                self.slack_service = SlackService(logger, token)


    def main(self):
        """
        Slackから投稿を取得

        Returns:

        """

        # 期間を取得
        oldest, latest = self.get_term()
        # チャンネルタイプの算出
        channel_type_list = []
        if int(self.conf['is_export_public']):
            channel_type_list.append(const.PUBLIC_CHANNEL)
        if int(self.conf['is_export_private']):
            channel_type_list.append(const.PRIVATE_CHANNEL)
        if int(self.conf['is_export_im']):
            channel_type_list.append(const.IM_CHANNEL)
        if int(self.conf['is_export_mpim']):
            channel_type_list.append(const.MPIM_CHANNEL)

        if int(self.conf['is_get_messages']):
            self.logger.info('get slack messages', is_print=True)

            cursor = self.conn.cursor()
            # トランザクションを開始
            self.conn.execute('BEGIN TRANSACTION')

            try:
                # Slackから投稿内容を取得
                self._get_slack_messages(cursor, channel_type_list, oldest, latest)

                # コミット
                self.conn.commit()

            except sqlite3.Error as e:
                self.conn.rollback()
                raise e

        # Excel出力
        self.logger.info('export slack messages', is_print=True)
        self._create_slack_messages()

        # 設定ファイル更新
        # from_dateを当日に更新する
        if int(self.conf['is_get_messages']) and int(self.conf['is_overwrite_from_date']):
            json_data = app_shared_service.get_conf(self.root_dir)
            json_data['get_messages']['from_date'] = datetime.datetime.now().strftime('%Y/%m/%d')
            app_shared_service.write_conf(self.root_dir, const.SETTINGS_FILENAME, json_data)


    def _get_slack_messages(self, cursor, channel_type_list, oldest, latest):
        """
        Slackからチャンネルタイプごとに投稿内容を取得

        Args:
            cursor:
            channel_type_list:
            oldest:
            latest:

        Returns:

        """

        # 対象チャネルを取得
        # チャンネルの一覧を取得
        cond = [set_cond_model.Condition('channel_type', channel_type_list, 'in')]
        sort = [set_sort_model.OrderBy('channel_type'), set_sort_model.OrderBy('channel_name')]
        channel_list = self.channel_dataaccess.select(cond, sort)

        replace_emoji_list = []
        with open(os.path.join(self.root_dir, const.CONF_DIR, const.EMOJI_FILENAME)) as f:
            for data in f.readlines():
                data = data.replace('\n', '')
                replace_emoji_list.append({
                    'before': data.split()[0],
                    'after': chr(int(data.split()[1], 0)),
                })

        # ユーザー
        # メンションがslack_user_idになっているのでdisplay_nameに置換するためのリスト
        slack_user_list = self.slack_user_dataaccess.select_all()
        replace_user_list = []
        for slack_user in slack_user_list:
            replace_user_list.append({
                'before': slack_user.slack_user_id,
                'after': slack_user.user_id,
            })

        history_list = []
        for channel in channel_list:
            # チャンネル単位で投稿内容、返信内容を取得
            if self.is_json:
                history_list_ = self.get_channel_messages_json(channel.channel_id, channel.channel_name, replace_emoji_list, replace_user_list)
            else:
                history_list_ = self.get_channel_messages(channel.channel_id, oldest, latest, replace_emoji_list, replace_user_list)
            history_list.extend(history_list_)

        # 投稿内容、返信内容の登録
        dataaccess = SlackExportDataaccess(cursor)
        for history_info in history_list:
            history_params = {
                'channel_id': history_info['channel_id'],
                'post_date': history_info['post_date'],
                'post_slack_user_id': history_info['post_user'],
                'post_message': history_info['post_message']
            }
            dataaccess.upsert_history(history_params)
            # 登録されたID
            if cursor.lastrowid:
                channel_history_id = cursor.lastrowid
            else:
                # 既存のIDを取得
                channel_history_id = self.get_channel_history_id_by_logical_pk(history_info['post_date'])

            # 返信内容
            for reply_info in history_info['reply_list']:
                reply_params = {
                    'channel_history_id': channel_history_id,
                    'reply_date': reply_info['reply_date'],
                    'reply_slack_user_id': reply_info['reply_user'],
                    'reply_message': reply_info['reply_message']
                }
                dataaccess.upsert_reply(reply_params)


    def get_channel_messages(self, channel_id, oldest, latest, replace_emoji_list, replace_user_list):
        """
        Slackのチャネル内の投稿内容と返信内容を取得

        Args:
            channel_id:
            oldest:
            latest:

        Returns:

        """
        history_list = []
        # 投稿一覧を取得
        time.sleep(1)
        result_history = self.slack_service.get_history(channel_id, limit=1000, oldest=oldest, latest=latest)

        if not result_history['ok']:
            self.logger.error('get history error:' + result_history['error'], is_print=True)
            return

        conversation_history = result_history['messages']

        while True:
            # 返信内容を取得
            for data_history in conversation_history:
                history_record = {}
                # slackから取得した情報を追加
                try:
                    post_date = datetime.datetime.fromtimestamp(int(str(data_history['ts']).split('.')[0]))
                    history_record = {
                        'channel_id': channel_id,
                        'post_date':post_date,
                        'post_user': data_history['user'],
                        'post_message': _convert_message(data_history['text'], replace_emoji_list, replace_user_list)
                    }
                    # history_list.append({'channel_id': channel_id, 'post_date':float(data_history['ts']), 'post_user': data_history['user'], 'post_message': data_history['text']})
                except Exception as e:
                    shared_service.print_except(e, self.logger)
                    self.logger.debug(','.join(f'{key}:{value}' for key, value in data_history.items()))
                    pass

                # 返信内容取得
                # 1分間のリクエスト数に制限があるため待機
                time.sleep(1.2)
                result_replies = self.slack_service.get_replies(channel_id, data_history['ts'])

                if not result_replies['ok']:
                    self.logger.error('get history error:' + result_replies['error'], is_print=True)
                    self.logger.debug(','.join(f'{key}:{value}' for key, value in data_history.items()))
                    break

                conversation_replies = result_replies['messages']

                reply_list = []
                for data_replies in conversation_replies:
                    if data_history['ts'] == data_replies['ts']:
                        # 返信内容には投稿自体の情報も含まれるため、投稿自体の情報は省く
                        continue

                    # slackから取得した情報を追加
                    try:
                        reply_date = datetime.datetime.fromtimestamp(int(str(data_replies['ts']).split('.')[0]))
                        reply_list.append({
                            'reply_date': reply_date,
                            'reply_user': data_replies['user'],
                            'reply_message': _convert_message(data_replies['text'], replace_emoji_list, replace_user_list)
                        })
                    except Exception as e:
                        shared_service.print_except(e, self.logger)
                        self.logger.debug(','.join(f'{key}:{value}' for key, value in data_replies.items()))
                        pass

                history_record['reply_list'] = reply_list
                history_list.append(history_record)

            if not result_history['has_more']:
                break

            # 次の投稿一覧を取得
            result_history = self.slack_service.get_history_cursor(channel_id, result_history['response_metadata']['next_cursor'])

            if not result_history['ok']:
                self.logger.error('get history error:' + result_history['error'], is_print=True)
                break

            conversation_history = result_history['messages']

        return history_list


    def get_channel_messages_json(self, channel_id, channel_name, replace_emoji_list, replace_user_list):
        """
        Slackのチャネル内の投稿内容と返信内容を取得

        Args:
            channel_id:
            channel_name:

        Returns:

        """
        base_dir = os.path.join(self.root_dir, 'import', 'get_messages')
        message_dir = os.path.join(base_dir, channel_name)
        if not os.path.isdir(message_dir):
            return []

        all_data = []
        # jsonファイルを読み込む
        for file_name in os.listdir(message_dir):
            if file_name.endswith('.json'):
                file_path = os.path.join(message_dir, file_name)
                with open(file_path, 'r', encoding='utf-8') as f:
                    json_data = json.load(f)
                    for data in json_data:
                        all_data.append(data)

        history_list = []
        for data in all_data:
            # 親スレッドかどうか
            if not 'ts' in data:
                continue
            if not 'thread_ts' in data or data['ts'] == data['thread_ts']:
                post_date = datetime.datetime.fromtimestamp(int(str(data['ts']).split('.')[0]))
                history_record = {
                    'channel_id': channel_id,
                    'post_date': post_date.strftime('%Y-%m-%d %H:%M:%S'),
                    'post_user': data['user'],
                    'post_message': _convert_message(data['text'], replace_emoji_list, replace_user_list),
                }

                # 返信内容を取得
                reply_list = []
                if not 'replies' in data:
                    history_record['reply_list'] = []
                else:
                    for replies in data['replies']:
                        for data2 in all_data:
                            if data2['ts'] == replies['ts']:
                                reply_date = datetime.datetime.fromtimestamp(int(str(replies['ts']).split('.')[0]))
                                reply_list.append({
                                    'reply_date': reply_date.strftime('%Y-%m-%d %H:%M:%S'),
                                    'reply_user': data2['user'],
                                    'reply_message': _convert_message(data2['text'], replace_emoji_list, replace_user_list),
                                })
                                break
                    history_record['reply_list'] = reply_list

                history_list.append(history_record)

        return history_list

    def get_term(self):
        """
        投稿内容取得期間を取得

        Returns:

        """
        from_date = self.conf['from_date']
        to_date = self.conf['to_date']
        if not from_date:
            # Fromが空の場合は1週間前を指定する
            from_date = (datetime.datetime.now() - datetime.timedelta(days=7)).strftime('%Y/%m/%d')
        if not to_date:
            # Toが空の場合は当日を指定する
            to_date = datetime.datetime.now().strftime('%Y/%m/%d')

        self.logger.info(f'get_messages : {from_date} - {to_date}', is_print=True)

        # UNIX時間に変換
        oldest = datetime.datetime.strptime(from_date, '%Y/%m/%d').timestamp()
        latest = datetime.datetime.strptime(to_date, '%Y/%m/%d').timestamp()

        return oldest, latest


    def _get_channel_list(self, channel_type_list=None):
        """
        チャンネル一覧を取得

        Args:
            channel_type_list:

        Returns:

        """
        sort = [
            set_sort_model.OrderBy('channel_type'),
            set_sort_model.OrderBy('channel_name')
        ]
        if channel_type_list:
            cond = [
                set_cond_model.Condition('channel_tyype', channel_type_list, 'in')
            ]
            return self.channel_dataaccess.select(conditions=cond, order_by_list=sort)

        return self.channel_dataaccess.select_all(order_by_list=sort)


    def _create_slack_messages(self):
        """
        Slackから取得した投稿内容を出力する

        Args:

        Returns:

        """
        # チャンネル一覧を取得
        channel_list = self._get_channel_list()

        # チャンネルごとにExcelファイルを生成
        for channel in channel_list:
            self.logger.info(channel.channel_name, is_print=True)

            # 投稿内容/返信内容を取得
            message_list = self._get_message_list(channel.channel_type, channel.channel_name)

            # 結果をExcelに出力
            output_dir = os.path.join(self.root_dir, const.EXPORT_DIR, channel.channel_type)
            os.makedirs(output_dir, exist_ok=True)
            output_file = os.path.join(output_dir, f'{channel.channel_name}.xlsx')
            pd.DataFrame(message_list).to_excel(output_file, sheet_name=channel.channel_name, index=False)

            # Excelフォーマット
            self._format(output_file)


    def _get_message_list(self, channel_type, channel_name) -> list | None:
        """
        投稿内容/返信内容を取得

        Args:
            channel_type:
            channel_name:

        Returns:

        """
        # データを取得
        cursor = self.conn.cursor()
        dataaccess = SlackExportDataaccess(cursor)
        results = dataaccess.get_slack_messages({'channel_type': channel_type, 'channel_name': channel_name})

        if results.empty:
            # データがない場合は出力しない
            return None

        # Excel出力用に整形
        message_list = []
        history_id_list = []
        no = 1
        for _, row in results.iterrows():
            message_list.append({
                'no': no,
                'post_icon': '' if row['channel_history_id'] in history_id_list else row['post_name'],
                'post_name': row['post_name'],
                'post_date': row['post_date'],
                'post_message': '' if row['channel_history_id'] in history_id_list else row['post_message'],
                'reply_icon': row['reply_name'],
                'reply_name': row['reply_name'],
                'reply_date': row['reply_date'],
                'reply_message': row['reply_message'],
            })
            if not row['channel_history_id'] in history_id_list:
                history_id_list.append(row['channel_history_id'])
                no += 1

        return message_list


    def _format(self, output_file):
        """
        Excelフォーマット

        Args:
            output_file:

        Returns:

        """

        # 罫線
        side = Side(style='thin', color='000000')
        border_outline = Border(top=side, bottom=side, left=side, right=side)
        # ヘッダ背景色
        fill_header = PatternFill(patternType='solid', fgColor='9FE2BF')
        # フォント
        font_base = Font(name='Meiryo UI', size=9)

        # 列の設定定義
        # (width, text-align, vertical, wrap_text)
        column_settings = {
            1: ('no', 10, 'left', 'top', False),
            2: ('post_icon', 5, 'left', 'top', False),
            3: ('post_name', 15, 'left', 'top', False),
            4: ('post_date', 20, 'center', 'top', False),
            5: ('post_message', 60, 'left', 'top', True),
            6: ('reply_icon', 5, 'left', 'top', False),
            7: ('reply_name', 15, 'left', 'top', False),
            8: ('reply_date', 20, 'center', 'top', False),
            9: ('reply_message', 60, 'left', 'top', True),
        }

        # フォーマット
        wb = openpyxl.load_workbook(output_file)
        ws = wb.worksheets[0]

        # 枠線非表示
        ws.sheet_view.showGridLines = False

        # ヘッダのフォーマット
        for col in ws.iter_cols(min_col=1, max_col=len(column_settings), min_row=1, max_row=1):
            for cell in col:
                cell.font = font_base
                cell.fill = fill_header
                cell.border = border_outline

        # 明細行のフォーマット/データの整備
        for col_idx, (col_id, width, align, vertical, wrap_text) in column_settings.items():
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    # 書式設定
                    cell.font = font_base
                    cell.alignment = Alignment(horizontal=align, vertical=vertical, wrap_text=wrap_text)
                    cell.border = border_outline

                    if col_id == 'post_icon':
                        # 投稿者アイコン
                        if cell.value:
                            icon_name = str(cell.value)
                            img_path = os.path.join(app_shared_service.get_icon_dir(self.bin_dir), f'{icon_name}.jpg')
                            if os.path.exists(img_path):
                                img = Image(img_path)
                                ws.add_image(img, cell.coordinate)
                            cell.value = None

                    if col_id == 'reply_icon':
                        # 返信者アイコン
                        if cell.value:
                            icon_name = str(cell.value)
                            img_path = os.path.join(app_shared_service.get_icon_dir(self.bin_dir), f'{icon_name}.jpg')
                            if os.path.exists(img_path):
                                img = Image(img_path)
                                ws.add_image(img, cell.coordinate)
                            cell.value = None

        if ws.max_row > 1:
            # 条件付き書式
            border_top_none = Border(top=Side(border_style=None))
            border_top_hair = Border(top=Side(border_style='hair', color='000000'))
            empty_font = Font(color='FFFFFF')
            formula = '=INDIRECT(ADDRESS(ROW(), 5)) = ""'

            rule1 = FormulaRule(formula=[formula], font=empty_font, border=border_top_none)
            ws.conditional_formatting.add(f'A2:E{ws.max_row}', rule1)

            rule2 = FormulaRule(formula=[formula], border=border_top_hair)
            ws.conditional_formatting.add(f'F2:I{ws.max_row}', rule2)

        # オートフィルタ
        ws.auto_filter.ref = ws.dimensions
        # ウィンドウ枠の固定
        ws.freeze_panes = 'B2'

        # 保存
        wb.save(output_file)


    def get_channel_history_id_by_logical_pk(self, post_date):
        """
        論理キーで投稿履歴IDを取得する

        Args:
            post_date:

        Returns:

        """
        cond = [set_cond_model.Condition('post_date', post_date)]
        results = self.tr_channel_histories_dataaccess.select(conditions=cond)

        return results[0].channel_id


def get_column_letter(n):
    """
    カラムインデックスから列名を取得

    Args:
        n:

    Returns:

    """
    result = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _convert_message(val, replace_emoji_list, replace_user_list):
    """
    絵文字、ユーザー名を置換

    Args:
        val:
        replace_emoji_list:
        replace_user_list:

    Returns:

    """
    if not val:
        return ''
    result = val
    for user in replace_user_list:
        if user['before'] in result:
            result = result.replace(user['before'], user['after'])
    for emoji in replace_emoji_list:
        if emoji['before'] in result:
            result = result.replace(emoji['before'], emoji['after'])

    return result
